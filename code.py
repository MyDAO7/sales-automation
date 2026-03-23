
import pandas as pd
import numpy as np
from  openpyxl  import Workbook,load_workbook
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================
# LOAD DATA
# ============================================
print("="*60)
print("CLIENT DATA ANALYSIS")
print("="*60)

# Load the client's file
df = pd.read_excel('client_sales_data.xlsx')

print(f"\n📊 Original data shape: {df.shape}")
print(f"\n📋 First 5 rows:")
print(df.head())

print(f"\n🔍 Missing values:")
print(df.isnull().sum())

print(f"\n🏷️ Unique branches: {df['Branch'].unique()}")

print(f"\n🔄 Duplicate rows: {df.duplicated().sum()}")

# ============================================
# DATA CLEANING
# ============================================
print("\n" + "="*60)
print("DATA CLEANING")
print("="*60)

# Fix missing amounts - fill with product average
df['Amount'] = df['Amount'].fillna(df.groupby('Product')['Amount'].transform('mean'))
print(f"\n✅ Missing amounts filled: {df['Amount'].isna().sum()} remaining")

# Fix missing dates
df['Date_Filled'] = df['Date'].isna()
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df['Date'] = df['Date'].fillna(pd.Timestamp('2024-01-01'))
print(f"✅ Missing dates filled: {df['Date_Filled'].sum()}")

# Fix branch typos
valid_branches = ['North', 'South', 'East']
df['Branch'] = df['Branch'].replace('West', 'East')
invalid_branches = df[~df['Branch'].isin(valid_branches)]
if len(invalid_branches) > 0:
    print(f"⚠️ Invalid branches found: {invalid_branches['Branch'].unique()}")
    df['Branch'] = df['Branch'].replace(invalid_branches['Branch'].unique(), 'Unknown')
else:
    print("✅ All branches are valid")

print(f"\nFixed branches: {df['Branch'].unique()}")

# Remove duplicates
before = len(df)
df = df.drop_duplicates()
after = len(df)
print(f"✅ Removed {before - after} duplicate rows")
print(f"✅ After cleaning: {len(df)} rows")

# ============================================
# CREATE MONTH COLUMN (NEEDED FOR MONTHLY SUMMARY)
# ============================================
print("\n" + "="*60)
print("CREATING MONTH COLUMN")
print("="*60)

df['Month'] = pd.to_datetime(df['Date']).dt.to_period('M').astype(str)
print("✅ Month column created")
print(df[['Date', 'Month']].head())

# ============================================
# CREATE SUMMARIES
# ============================================
print("\n" + "="*60)
print("CREATING SUMMARIES")
print("="*60)

summarry = {}

# 1. By Branch (flatten MultiIndex)
branch_raw = df.pivot_table(
    index='Branch',
    values='Amount',
    aggfunc=['sum', 'count', 'mean']
).reset_index()

# Flatten column names
branch_raw.columns = ['Branch', 'Total_Sales', 'Order_Count', 'Avg_Sale']

# Add grand total
total_row = pd.DataFrame({
    'Branch': ['GRAND TOTAL'],
    'Total_Sales': [branch_raw['Total_Sales'].sum()],
    'Order_Count': [branch_raw['Order_Count'].sum()],
    'Avg_Sale': [branch_raw['Avg_Sale'].mean()]
})

summarry['By_Branch'] = pd.concat([branch_raw, total_row], ignore_index=True)
print("✅ Branch summary with grand total created")

# 2. Top Products
summarry['Top_product'] = df.groupby('Product')['Amount'].sum().sort_values(ascending=False).head(5).reset_index()
summarry['Top_product'].columns = ['Product', 'Total_Sales']
print("✅ Top products created")

# 3. Monthly Summary (flatten MultiIndex)
monthly_raw = df.pivot_table(
    index='Month',
    values='Amount',
    aggfunc=['sum', 'count', 'mean']
).reset_index()

monthly_raw.columns = ['Month', 'Total_Sales', 'Order_Count', 'Avg_Sale']

# Add grand total for monthly
total_month_row = pd.DataFrame({
    'Month': ['GRAND TOTAL'],
    'Total_Sales': [monthly_raw['Total_Sales'].sum()],
    'Order_Count': [monthly_raw['Order_Count'].sum()],
    'Avg_Sale': [monthly_raw['Avg_Sale'].mean()]
})

summarry['monthly'] = pd.concat([monthly_raw, total_month_row], ignore_index=True)
print("✅ Monthly summary with grand total created")

# 4. Overall Stats
stats_dict = {
    'Metric': [
        'Total Sales',
        'Total Orders',
        'Average Order Value',
        'Unique Products',
        'Unique Branches',
        'Date Range',
        'Missing Dates Fixed',
        'Duplicates Removed'
    ],
    'Value': [
        f"${df['Amount'].sum():,.2f}",
        len(df),
        f"${df['Amount'].mean():,.2f}",
        df['Product'].nunique(),
        df['Branch'].nunique(),
        f"{df['Date'].min().date()} to {df['Date'].max().date()}",
        df['Date_Filled'].sum(),
        before - after
    ]
}
summarry['Overall'] = pd.DataFrame(stats_dict)
print("✅ Overall stats created")

# ============================================
# CREATE LINE CHART
# ============================================
print("\n" + "="*60)
print("CREATING CHART")
print("="*60)

# Use monthly data without grand total for chart
chart_data = monthly_raw[monthly_raw['Month'] != 'GRAND TOTAL'].copy()

# Create line chart
plt.figure(figsize=(10, 6))
plt.plot(chart_data['Month'], chart_data['Total_Sales'], marker='o', linewidth=2, markersize=8, color='blue')
plt.title('Monthly Sales Trend', fontsize=14, fontweight='bold')
plt.xlabel('Month', fontsize=12)
plt.ylabel('Total Sales ($)', fontsize=12)
plt.grid(True, alpha=0.3)
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('monthly_trend.png', dpi=100, bbox_inches='tight')
plt.show()
plt.close()
print("✅ Chart saved: monthly_trend.png")

# ============================================
# SAVE FINAL EXCEL
# ============================================
print("\n" + "="*60)
print("SAVING FINAL REPORT")
print("="*60)

with pd.ExcelWriter('Sales_Report_Final.xlsx', engine='openpyxl') as writer:
    # Cleaned data
    df.to_excel(writer, sheet_name='Cleaned Data', index=False)
    print("✅ Added: Cleaned Data")
    
    # Overall stats
    summarry['Overall'].to_excel(writer, sheet_name='Overall Statistics', index=False)
    print("✅ Added: Overall Statistics")
    
    # Branch summary with grand total
    summarry['By_Branch'].to_excel(writer, sheet_name='Sales by Branch', index=False)
    print("✅ Added: Sales by Branch")
    
    # Monthly summary with grand total
    summarry['monthly'].to_excel(writer, sheet_name='Monthly Summary', index=False)
    print("✅ Added: Monthly Summary")
    
    # Top products
    summarry['Top_product'].to_excel(writer, sheet_name='Top Products', index=False)
    print("✅ Added: Top Products")

print(f"\n✅ Final report saved: Sales_Report_Final.xlsx")


monthly_data = summarry['monthly'][summarry['monthly']['Month'] != 'GRAND TOTAL'].copy()
monthly_data.plot(kind='line',x='Month',y='Total_Sales',figsize=(6,4),marker='o')
plt.title('Monthly Sales Trend')
plt.xlabel('Month')
plt.ylabel('Total Sales ($)')
plt.grid(True)
plt.tight_layout()
plt.savefig('chart.png', dpi=100)
plt.close()
wb= load_workbook('Sales_Report_Final.xlsx')
ws=wb['Monthly Summary']
img=Image('chart.png')
img.width=400
img.height=250
ws.add_image(img,'F2')
wb.save('Sales_Report_Final.xlsx')

print("✅ Chart added!")
branch_data = summarry['By_Branch'][summarry['By_Branch']['Branch'] != 'GRAND TOTAL'].copy()

# Define professional colors
colors = ['#366092', '#4F81BD', '#8FAADC', '#B8CCE4']  # Blue shades
# Or use: colors = ['gold', 'lightcoral', 'lightskyblue', 'lightgreen']

# Create pie chart
plt.figure(figsize=(8, 8))
branch_data.set_index('Branch')['Total_Sales'].plot(
    kind='pie',
    autopct='%1.1f%%',      # Percentage format
    colors=colors,           # Custom colors
    startangle=90,           # Rotate chart (optional)
    shadow=True,             # Add shadow (optional)
    explode=(0.05, 0, 0), # Slightly separate first slice
    title='Sales by Branch'
)
plt.ylabel('')  # Remove y-label
plt.tight_layout()
plt.savefig('branch_pie.png', dpi=100)
plt.close()

# Add to Excel
wb = load_workbook('Sales_Report_Final.xlsx')
ws = wb['Sales by Branch']
img = Image('branch_pie.png')
img.width = 450
img.height = 400
ws.add_image(img, 'E2')
wb.save('Sales_Report_Final.xlsx')

print("✅ Colored pie chart added!")
# ============================================
# FINAL SUMMARY
# ============================================
print("\n" + "="*60)
print("PROJECT COMPLETE")
print("="*60)
print(f"""
📊 SUMMARY:
   - Original rows: {before + (before - after)}
   - Cleaned rows: {len(df)}
   - Duplicates removed: {before - after}
   - Missing dates fixed: {df['Date_Filled'].sum()}
   - Total sales: ${df['Amount'].sum():,.2f}
   - Best branch: {branch_raw.iloc[0]['Branch']} (${branch_raw.iloc[0]['Total_Sales']:,.2f})
   - Top product: {summarry['Top_product'].iloc[0]['Product']} (${summarry['Top_product'].iloc[0]['Total_Sales']:,.2f})

📁 FILES CREATED:
   - Sales_Report_Final.xlsx (5 sheets)
   - monthly_trend.png (line chart)

📄 EXCEL SHEETS:
   1. Cleaned Data - All cleaned sales records
   2. Overall Statistics - Key metrics at a glance
   3. Sales by Branch - Sales per branch (with Grand Total)
   4. Monthly Summary - Sales trend by month (with Grand Total)
   5. Top Products - Best selling products

✅ Ready for client delivery!
""")

# Show first few rows of cleaned data for verification
print("\n" + "="*60)
print("SAMPLE OF CLEANED DATA")
print("="*60)
print(df.head(10).to_string())



# Load your Excel file
wb = load_workbook('Sales_Report_Final.xlsx')

# Define styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')

# For money columns
money_format = '#,##0.00'

# For number columns
number_format = '0'

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Skip chart sheets or sheets without headers
    if sheet_name in ['Sales Chart', 'Monthly Chart']:
        continue
    
    # Format headers (row 1)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Format data rows (row 2 onwards)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Format numbers in columns with sales
            if 'Sales' in str(ws.cell(row=1, column=col).value) or 'Total' in str(ws.cell(row=1, column=col).value):
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
            
            # Format count/order columns
            if 'Order' in str(ws.cell(row=1, column=col).value) or 'Count' in str(ws.cell(row=1, column=col).value):
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format
    
    # Auto-fit columns
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

# Save the formatted file
wb.save('Sales_Report_Final_Formatted.xlsx')
print("✅ Professional formatting added to all sheets!")
